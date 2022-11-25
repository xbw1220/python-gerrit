from tabulate import tabulate
import json, fileinput, logging
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from gerrit import Gerrit

Gerrit = Gerrit()
login = Gerrit.login()
user_dict = {
    "郑志安": "zhengzhian_zh",
    # "黎伟超": "liweichao_zh",
    # "林国权": "linguoquan_zh",
    # "梁惜家": "liangxijia_zh",
    # "冯海": "fenghai_zh",
    # "廖飞亮": "liaofeiliang_zh",
    # "段金亮": "duanjinliang_zh",
    # "黄国印": "huangguoying_zh",
    # "梁海明": "lianghaiming_zh",
    # "李化武": "lihuawu_zh",
    # "李杰": "lijie_zh",
    # "刘泽增": "liuzezeng_zh",
    # "龙宪彪": "longxianbiao_zh",
    # "卢继瑶": "lujiyao_zh",
    # "庞智明": "pangzhiming_zh",
    # "苏品林": "supinlin_zh",
    # "吴华君": "wuhuajun_zh",
    # "杨邦淳": "yangbangchun_zh",
    # "闫兰松": "yanlansong_zh",
    # "曾庆凡": "zengqingfan_zh",
    # "赵德成": "zhaodecheng_zh",
}
start_time = '2022-07-01'
end_time = '2022-11-23'
datas = []
# print(type(user_dict))
for chinese_name in user_dict:
    username = user_dict[chinese_name]
    start_num = 0
    # print(Gerrit.query_para('yufan0528', '2022-01-01', '2022-11-23', 0))
    query = Gerrit.query_change_para(username, start_time, end_time)
    print("/changes/?q=%s&start=%d" % ("%20".join(query), start_num))
    query_changes = login.get("/changes/?q=%s&start=%d&o=CURRENT_REVISION" % ("%20".join(query), start_num))
    for change in query_changes:
        try:
            if change["_more_changes"] == True:
                start_num += 500
                query_changes += login.get(
                    "/changes/?q=%s&start=%d&o=CURRENT_REVISION" % ("%20".join(query), start_num))
        except Exception:
            pass
            # print(change)
            # print(query_changes.index(change))
            # print("id: %s, change_id: %s,revision_id: %s, project:%s, branch: %s" % (
            #     change["id"], change["change_id"], change["current_revision"], change["project"], change["branch"]))
            # logging.info("id: %s, change_id: %s, project:%s, branch: %s", change["id"], change["change_id"],
            #              change["project"], change["branch"])

    insertions, deletions = 0, 0
    for change in query_changes:
        insertions += change["insertions"]
        deletions += change["deletions"]
    change_files_num = 0
    ids_list = []
    for change in query_changes:
        id_tmp = []
        id_tmp.append(change["id"])
        id_tmp.append(change["current_revision"])
        ids_list.append(id_tmp)
    # print(ids_list)
    for id, revision_id in ids_list:
        query_change_files = login.get("/changes/%s/revisions/%s/files/" % (id, revision_id))
        query_change_files.pop("/COMMIT_MSG")
        try:
            query_change_files.pop(".gitignore")
        except Exception:
            pass
        change_files_num += len(query_changes)
        # print(type(query_change_files))
    owner_data = []
    owner_data.append(chinese_name)
    owner_data.append(username)
    owner_data.append(start_time + '~' + end_time)
    owner_data.append(len(query_changes))
    owner_data.append(change_files_num)
    owner_data.append("+" + str(insertions) + "," + "-" + str(deletions))
    # print(owner_data)
    datas.append(owner_data)

print(datas)
# wb = load_workbook('./data.xlsx')
# ws = wb.active
# cell = ws['A1']
# for i in datas:
#     ws.append(i)
# wb.save(filename='./data.xlsx')

wb = Workbook()
ws = wb.active
font = Font(
    name="等线",  # 字体
    size=11,  # 字体大小
    color="000000",  # 字体颜色，用16进制rgb表示
    bold=False,  # 是否加粗，True/False
    italic=False,  # 是否斜体，True/False
    strike=None,  # 是否使用删除线，True/False
    underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
)
cell = ws['A1']
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 11
ws.column_dimensions["E"].width = 20
ws.column_dimensions["F"].width = 20
ws['A1'] = '提交人'
ws['A1'].font = font
ws['B1'] = 'Gerrit账号'
ws['B1'].font = font
ws['C1'] = '提交时间'
ws['C1'].font = font
ws['D1'] = '提交次数'
ws['D1'].font = font
ws['E1'] = '提交修改文件总数'
ws['E1'].font = font
ws['F1'] = '增删行总数'
ws['F1'].font = font

for i in datas:
    ws.append(i)

ws.font = font
wb.save(filename='./test.xlsx')
wb.close()
