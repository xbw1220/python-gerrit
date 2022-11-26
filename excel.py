import config
from gerrit import Gerrit
from openpyxl.styles import Font
from openpyxl import load_workbook, Workbook

Gerrit = Gerrit()
login = Gerrit.login()


def get_total_data(start_time: str, end_time: str, users: dict):
    datas = []
    for chinese_name in users:
        username = users[chinese_name]
        print(username)
        start_num = 0
        # print(Gerrit.query_para('yufan0528', '2022-01-01', '2022-11-23', 0))
        query = Gerrit.query_change_para(username, start_time, end_time)
        print("/changes/?q=%s&start=%d" % ("%20".join(query), start_num))
        query_changes = login.get("/changes/?q=%s&start=%d&o=CURRENT_REVISION" % ("%20".join(query), start_num))
        for change in query_changes:
            try:
                if change["_more_changes"] == True:
                    start_num += 500
                    query_changes += login.get(
                        "/changes/?q=%s&start=%d&o=CURRENT_REVISION" % ("%20".join(query), start_num))
            except Exception:
                pass
                # print(change)
                # print(query_changes.index(change))
                # print("id: %s, change_id: %s,revision_id: %s, project:%s, branch: %s" % (
                #     change["id"], change["change_id"], change["current_revision"], change["project"], change["branch"]))
                # logging.info("id: %s, change_id: %s, project:%s, branch: %s", change["id"], change["change_id"],
                #              change["project"], change["branch"])

        insertions, deletions = 0, 0
        for change in query_changes:
            insertions += change["insertions"]
            deletions += change["deletions"]
        change_files_num = 0
        ids_list = []
        for change in query_changes:
            id_tmp = []
            id_tmp.append(change["id"])
            id_tmp.append(change["current_revision"])
            ids_list.append(id_tmp)
        # print(ids_list)
        for id, revision_id in ids_list:
            query_change_files = login.get("/changes/%s/revisions/%s/files/" % (id, revision_id))
            query_change_files.pop("/COMMIT_MSG")
            try:
                query_change_files.pop(".gitignore")
            except Exception:
                pass
            change_files_num += len(query_changes)
            # print(type(query_change_files))
        owner_data = []
        owner_data.append(chinese_name)
        owner_data.append(username)
        owner_data.append(start_time + '~' + end_time)
        owner_data.append(len(query_changes))
        owner_data.append(change_files_num)
        owner_data.append("+" + str(insertions) + "," + "-" + str(deletions))
        # print(owner_data)
        datas.append(owner_data)
    return datas


def create_total_excel_file(datas: list, excel_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = '总统计数据'
    font = Font(
        name="等线",  # 字体
        size=11,  # 字体大小
        color="000000",  # 字体颜色，用16进制rgb表示
        bold=False,  # 是否加粗，True/False
        italic=False,  # 是否斜体，True/False
        strike=None,  # 是否使用删除线，True/False
        underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
    )
    # cell = ws['A1']
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws['A1'] = '提交人'
    ws['A1'].font = font
    ws['B1'] = 'Gerrit账号'
    ws['B1'].font = font
    ws['C1'] = '提交时间'
    ws['C1'].font = font
    ws['D1'] = '提交次数'
    ws['D1'].font = font
    ws['E1'] = '提交修改文件总数'
    ws['E1'].font = font
    ws['F1'] = '增删行总数'
    ws['F1'].font = font
    ws.auto_filter.ref = "A1:F1"
    for i in datas:
        ws.append(i)
    ws.font = font
    wb.save(filename=excel_name)
    wb.close()
    print('excel统计表创建完成')


def update_excel_file(user: str, user_file_list: list, excel_name: str):
    wb = load_workbook(filename=excel_name)
    wb.create_sheet(title=user)

    sheet = wb[user]
    sheet.column_dimensions["A"].width = 100
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 10
    sheet['A1'] = '修改文件'
    sheet['B1'] = '是否为二进制文件'
    sheet['C1'] = '新增行数'
    sheet['D1'] = '删减行数'
    sheet.auto_filter.ref = "A1:D1"
    for i in user_file_list:
        sheet.append(i)
    wb.save(filename=excel_name)
    wb.close()
    print('excel统计表更新完成')


def get_user_data(start_time: str, end_time: str, users: dict, excel_name: str):
    for chinese_name in users:
        user_file_list = []
        username = users[chinese_name]
        print(username)
        start_num = 0
        query = Gerrit.query_change_para(username, start_time, end_time)
        # print("/changes/?q=%s&start=%d" % ("%20".join(query), start_num))
        query_changes = login.get("/changes/?q=%s&start=%d&o=CURRENT_REVISION" % ("%20".join(query), start_num))
        for change in query_changes:
            try:
                if change["_more_changes"] == True:
                    start_num += 500
                    query_changes += login.get(
                        "/changes/?q=%s&start=%d&o=CURRENT_REVISION" % ("%20".join(query), start_num))
            except Exception:
                pass

        ids_list = []
        for change in query_changes:
            id_tmp = []
            id_tmp.append(change["id"])
            id_tmp.append(change["current_revision"])
            ids_list.append(id_tmp)
        for id, revision_id in ids_list:
            query_change_files = login.get("/changes/%s/revisions/%s/files/" % (id, revision_id))
            query_change_files.pop("/COMMIT_MSG")
            try:
                # pass
                query_change_files.pop(".gitignore")
            except Exception:
                pass
            # print(query_change_files)
            tmp_list = []
            for filename in query_change_files:
                print(filename)
                # print('1-' * 100)
                # print(query_change_files[filename])
                # print('2-' * 100)
                tmp_list.append(filename)
                # 是否是二进制
                try:
                    if query_change_files[filename]['binary'] == True:
                        tmp_list.append("Yes")
                        break
                except Exception:
                    tmp_list.append("No")
                # 是否有新增行
                try:
                    insert_line = query_change_files[filename]['lines_inserted']
                    tmp_list.append('+' + str(insert_line))
                except Exception:
                    tmp_list.append(None)
                # 是否有删减行
                try:
                    delete_line = query_change_files[filename]['lines_deleted']
                    tmp_list.append('-' + str(delete_line))
                except Exception:
                    tmp_list.append(None)
                break
                # print(tmp_list)
                # print('3-' * 100)
            user_file_list.append(tmp_list)
        # print(user_file_list)
        update_excel_file(chinese_name, user_file_list, excel_name)


# start_time = '2022-07-01'
# end_time = '2022-11-23'
# get_user_data(start_time, end_time, config.user_dict)
