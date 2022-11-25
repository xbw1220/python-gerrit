from openpyxl import load_workbook

wb = load_workbook('./data.xlsx')
print(wb)

ws = wb.active  # 获取当前活跃的worksheet对象（sheet表）
print(ws)

cell = ws['A1']  # 获取指定位置的单元格对象
print(cell)

ws.append([1, 2, 3])

wb.save(filename='./data.xlsx')